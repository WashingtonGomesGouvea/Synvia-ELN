"""
📊 Dashboard Synvia ELN - Controle de Qualidade
Monitoramento de estudos de bioequivalência/biodisponibilidade
Execute com: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from auth_microsoft import (
    MicrosoftAuth,
    AuthManager,
    create_login_page,
    create_user_header,
    LOGIN_CONFIG
)
from sp_connector import SPConnector

# ============================================================================
# CONFIGURAÇÃO DA PÁGINA DE LOGIN
# ============================================================================
LOGIN_CONFIG.update({
    "title": "Synvia ELN Dashboard",
    "subtitle": "Sistema de Controle de Qualidade - Bioequivalência",
    "badge_text": "Acesso Restrito",
    "email_domain": "@synvia.com.br",
    "highlights": [
        {
            "icon": "🔬",
            "title": "Controle de Qualidade",
            "description": "Monitoramento em tempo real dos estudos"
        },
        {
            "icon": "📊",
            "title": "Dashboard Interativa",
            "description": "KPIs, gráficos e filtros avançados"
        },
        {
            "icon": "🏭",
            "title": "28 Empresas",
            "description": "Acompanhamento de todos os clientes"
        }
    ]
})

# ============================================================================
# CONFIGURAÇÃO DA PÁGINA STREAMLIT
# ============================================================================
st.set_page_config(
    page_title="Synvia ELN Dashboard",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# CSS CUSTOMIZADO
# ============================================================================
st.markdown("""
<style>
    /* Cards de KPI */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    div[data-testid="stMetric"] label {
        color: rgba(255,255,255,0.8) !important;
        font-size: 0.9rem !important;
    }
    
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 2rem !important;
        font-weight: 700 !important;
    }
    
    /* Header principal */
    .main-header {
        background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
        padding: 30px;
        border-radius: 16px;
        margin-bottom: 30px;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.3);
    }
    
    .main-header h1 {
        color: white;
        font-size: 2.5rem;
        margin: 0;
    }
    
    .main-header p {
        color: rgba(255,255,255,0.7);
        font-size: 1rem;
        margin-top: 10px;
    }
    
    /* Tabela estilizada */
    .styled-table {
        border-radius: 10px;
        overflow: hidden;
    }
    
    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# AUTENTICAÇÃO
# ============================================================================
auth = MicrosoftAuth()

if not create_login_page(auth, LOGIN_CONFIG):
    st.stop()

create_user_header()
AuthManager.check_and_refresh_token(auth)

user = AuthManager.get_current_user()
user_name = user.get("displayName", "Usuário") if user else "Usuário"

# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

@st.cache_data(ttl=300)  # Cache de 5 minutos
def load_data():
    """Carrega e processa os dados do CSV (SharePoint ou local)"""
    df = None
    source = "local"
    
    # Tentar carregar do SharePoint primeiro
    try:
        # Verificar se as credenciais do SharePoint estão configuradas (seção [graph])
        if "graph" in st.secrets:
            graph_cfg = st.secrets["graph"]
            sp = SPConnector(
                tenant_id=graph_cfg["tenant_id"],
                client_id=graph_cfg["client_id"],
                client_secret=graph_cfg["client_secret"],
                user_upn=graph_cfg.get("user_upn", "washington.gouvea@synvia.com")
            )
            
            # Caminho do arquivo no OneDrive
            file_path = graph_cfg.get(
                "file_path", 
                "Data Analysis/Demandas/Desenvolvimento/Streamlit Projects/Synvia ELN (BQV)/resumo_powerbi.csv"
            )
            
            df = sp.read_csv(file_path, sep=";", encoding="utf-8-sig", decimal=".")
            source = "sharepoint"
    except Exception as e:
        st.warning(f"⚠️ Não foi possível conectar ao SharePoint: {str(e)[:100]}. Usando arquivo local.")
        df = None
    
    # Fallback: carregar arquivo local
    if df is None:
        try:
            df = pd.read_csv(
                "resumo_powerbi.csv",
                sep=";",
                encoding="utf-8-sig",
                decimal="."
            )
            source = "local"
        except Exception as e:
            st.error(f"❌ Erro ao carregar dados: {e}")
            return pd.DataFrame(), "erro"
    
    # Processar dados
    # Extrair ano do estudo (ex: 001.001.24 → 2024)
    df['ANO'] = df['ESTUDO'].apply(lambda x: f"20{x.split('.')[-1]}" if pd.notna(x) else None)
    
    # Extrair nome da empresa (ex: 001_EMS → EMS)
    df['EMPRESA_NOME'] = df['EMPRESA'].apply(lambda x: x.split('_')[1] if pd.notna(x) and '_' in x else x)
    
    # Converter data
    df['ULTIMA_ATUALIZACAO'] = pd.to_datetime(df['ULTIMA_ATUALIZACAO'], format='%d/%m/%Y %H:%M', errors='coerce')
    
    return df, source


def create_kpi_card(label, value, delta=None, delta_color="normal"):
    """Cria um card de KPI estilizado"""
    st.metric(label=label, value=value, delta=delta, delta_color=delta_color)


def get_status_color(value, thresholds=(70, 90)):
    """Retorna cor baseada no valor"""
    if value >= thresholds[1]:
        return "#00d4aa"  # Verde
    elif value >= thresholds[0]:
        return "#ffd93d"  # Amarelo
    else:
        return "#ff6b6b"  # Vermelho


# ============================================================================
# CARREGAR DADOS
# ============================================================================
df, data_source = load_data()

if df.empty:
    st.error("❌ Não foi possível carregar os dados. Verifique se o arquivo 'resumo_powerbi.csv' existe.")
    st.stop()

# Indicador de fonte de dados na sidebar
if data_source == "sharepoint":
    st.sidebar.success("☁️ Dados: SharePoint")
else:
    st.sidebar.info("📁 Dados: Arquivo local")

# ============================================================================
# SIDEBAR - FILTROS
# ============================================================================
st.sidebar.markdown("---")
st.sidebar.markdown("## 🎛️ Filtros")

# Filtro de Empresa
empresas = ["Todas"] + sorted(df['EMPRESA_NOME'].dropna().unique().tolist())
empresa_selecionada = st.sidebar.selectbox("🏭 Empresa", empresas)

# Filtro de Ano
anos = ["Todos"] + sorted(df['ANO'].dropna().unique().tolist(), reverse=True)
ano_selecionado = st.sidebar.selectbox("📅 Ano do Estudo", anos)

# Filtro de Estudo (dinâmico baseado na empresa selecionada)
df_temp = df.copy()
if empresa_selecionada != "Todas":
    df_temp = df_temp[df_temp['EMPRESA_NOME'] == empresa_selecionada]
estudos = ["Todos"] + sorted(df_temp['ESTUDO'].dropna().unique().tolist())
estudo_selecionado = st.sidebar.selectbox("📋 Estudo", estudos)

# Filtro de Completude
completude_min = st.sidebar.slider(
    "📊 Completude Mínima (%)",
    min_value=0,
    max_value=100,
    value=0,
    step=5
)

# Checkbox para estudos críticos
apenas_criticos = st.sidebar.checkbox("⚠️ Apenas estudos críticos", value=False)

st.sidebar.markdown("---")

# Verificar última atualização e alertar se desatualizado
ultima_atualizacao = df['ULTIMA_ATUALIZACAO'].max() if not df['ULTIMA_ATUALIZACAO'].isna().all() else None

if ultima_atualizacao:
    dias_desde_atualizacao = (datetime.now() - ultima_atualizacao).days
    data_formatada = ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
    
    st.sidebar.markdown(f"📅 **Última atualização:** {data_formatada}")
    
    # Alerta se dados estiverem desatualizados (> 7 dias)
    if dias_desde_atualizacao > 7:
        st.sidebar.error(f"⚠️ **ATENÇÃO:** Dados desatualizados há {dias_desde_atualizacao} dias!")
        st.sidebar.markdown("📧 Contate: **felipe.minari@synvia.com**")
    elif dias_desde_atualizacao > 3:
        st.sidebar.warning(f"⏰ Dados com {dias_desde_atualizacao} dias de atraso")
else:
    st.sidebar.markdown("📅 **Última atualização:** N/A")

# ============================================================================
# APLICAR FILTROS
# ============================================================================
df_filtered = df.copy()

if empresa_selecionada != "Todas":
    df_filtered = df_filtered[df_filtered['EMPRESA_NOME'] == empresa_selecionada]

if ano_selecionado != "Todos":
    df_filtered = df_filtered[df_filtered['ANO'] == ano_selecionado]

if estudo_selecionado != "Todos":
    df_filtered = df_filtered[df_filtered['ESTUDO'] == estudo_selecionado]

df_filtered = df_filtered[df_filtered['%_COMPLETUDE'] >= completude_min]

if apenas_criticos:
    df_filtered = df_filtered[
        (df_filtered['%_APROVADOS'] < 70) | 
        (df_filtered['%_COMPLETUDE'] < 80)
    ]

# ============================================================================
# HEADER PRINCIPAL
# ============================================================================
st.markdown(f"""
<div class="main-header">
    <h1>🔬 Dashboard de Controle de Qualidade</h1>
    <p>Monitoramento de Estudos de Bioequivalência | Olá, {user_name}!</p>
</div>
""", unsafe_allow_html=True)

# Banner de alerta se dados MUITO desatualizados (> 30 dias - crítico)
if ultima_atualizacao and dias_desde_atualizacao > 30:
    st.error(f"""
    ⚠️ **ATENÇÃO: Os dados estão desatualizados há {dias_desde_atualizacao} dias!**
    
    Última atualização: **{data_formatada}**
    
    📧 Entre em contato com **felipe.minari@synvia.com** (fonte de dados) para solicitar a atualização do arquivo.
    """)

# ============================================================================
# KPIs PRINCIPAIS
# ============================================================================
col1, col2, col3, col4, col5, col6 = st.columns(6)

total_lotes = df_filtered['TOTAL_LOTES'].sum()
total_aprovados = df_filtered['APROVADOS'].sum()
total_reprovados = df_filtered['REPROVADOS'].sum()
taxa_aprovacao = (total_aprovados / total_lotes * 100) if total_lotes > 0 else 0
taxa_reprovacao = (total_reprovados / total_lotes * 100) if total_lotes > 0 else 0
completude_media = df_filtered['%_COMPLETUDE'].mean() if len(df_filtered) > 0 else 0
estudos_unicos = df_filtered['ESTUDO'].nunique()
empresas_unicas = df_filtered['EMPRESA_NOME'].nunique()

with col1:
    create_kpi_card("📦 Total de Lotes", f"{total_lotes:,.0f}")

with col2:
    create_kpi_card("✅ Taxa Aprovação", f"{taxa_aprovacao:.1f}%")

with col3:
    create_kpi_card("❌ Taxa Reprovação", f"{taxa_reprovacao:.1f}%")

with col4:
    create_kpi_card("📊 Completude Média", f"{completude_media:.1f}%")

with col5:
    create_kpi_card("📋 Estudos", f"{estudos_unicos}")

with col6:
    create_kpi_card("🏭 Empresas", f"{empresas_unicas}")

st.markdown("<br>", unsafe_allow_html=True)

# ============================================================================
# GRÁFICOS - LINHA 1
# ============================================================================
col_left, col_right = st.columns(2)

with col_left:
    st.markdown("### 📊 Taxa de Aprovação por Empresa")
    
    # Agregar dados por empresa
    df_empresa = df_filtered.groupby('EMPRESA_NOME').agg({
        'APROVADOS': 'sum',
        'REPROVADOS': 'sum',
        'TOTAL_LOTES': 'sum'
    }).reset_index()
    
    df_empresa['TAXA_APROVACAO'] = (df_empresa['APROVADOS'] / df_empresa['TOTAL_LOTES'] * 100).fillna(0)
    df_empresa = df_empresa[df_empresa['TOTAL_LOTES'] > 0].sort_values('TAXA_APROVACAO', ascending=True)
    
    # Cores baseadas na taxa
    colors = [get_status_color(v) for v in df_empresa['TAXA_APROVACAO']]
    
    fig_aprovacao = go.Figure(go.Bar(
        x=df_empresa['TAXA_APROVACAO'],
        y=df_empresa['EMPRESA_NOME'],
        orientation='h',
        marker_color=colors,
        text=df_empresa['TAXA_APROVACAO'].apply(lambda x: f'{x:.1f}%'),
        textposition='outside',
        hovertemplate='<b>%{y}</b><br>Taxa: %{x:.1f}%<extra></extra>'
    ))
    
    fig_aprovacao.update_layout(
        height=400,
        xaxis_title="Taxa de Aprovação (%)",
        yaxis_title="",
        xaxis=dict(range=[0, 110]),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        margin=dict(l=0, r=50, t=20, b=40)
    )
    
    # Linha de meta (80%)
    fig_aprovacao.add_vline(x=80, line_dash="dash", line_color="#ffd93d", annotation_text="Meta 80%")
    
    st.plotly_chart(fig_aprovacao, use_container_width=True)

with col_right:
    st.markdown("### 🔄 Pipeline Analítico")
    
    # Calcular médias do pipeline
    pipeline_data = {
        'Etapa': ['Aliquotagem', 'Dopagem', 'Extração', 'Injeção', 'Brutos'],
        'Percentual': [
            df_filtered['%_ALIQUOTAGEM'].mean(),
            df_filtered['%_DOPAGEM'].mean(),
            df_filtered['%_EXTRACAO'].mean(),
            df_filtered['%_INJECAO'].mean(),
            df_filtered['%_BRUTOS'].mean()
        ]
    }
    df_pipeline = pd.DataFrame(pipeline_data)
    
    fig_pipeline = go.Figure(go.Funnel(
        y=df_pipeline['Etapa'],
        x=df_pipeline['Percentual'],
        textposition="inside",
        textinfo="value+percent initial",
        texttemplate="%{value:.1f}%",
        marker=dict(
            color=['#00d4aa', '#00b894', '#0984e3', '#6c5ce7', '#fd79a8']
        ),
        connector={"line": {"color": "royalblue", "dash": "solid", "width": 2}}
    ))
    
    fig_pipeline.update_layout(
        height=400,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        margin=dict(l=0, r=0, t=20, b=20)
    )
    
    st.plotly_chart(fig_pipeline, use_container_width=True)

# ============================================================================
# GRÁFICOS - LINHA 2
# ============================================================================
col_left2, col_right2 = st.columns(2)

with col_left2:
    st.markdown("### 📈 Completude por Empresa")
    
    df_completude = df_filtered.groupby('EMPRESA_NOME').agg({
        '%_COMPLETUDE': 'mean',
        'TOTAL_LOTES': 'sum'
    }).reset_index()
    
    df_completude = df_completude[df_completude['TOTAL_LOTES'] > 0].sort_values('%_COMPLETUDE', ascending=False).head(15)
    
    colors_completude = [get_status_color(v, (60, 80)) for v in df_completude['%_COMPLETUDE']]
    
    fig_completude = go.Figure()
    
    fig_completude.add_trace(go.Bar(
        x=df_completude['EMPRESA_NOME'],
        y=df_completude['%_COMPLETUDE'],
        marker_color=colors_completude,
        text=df_completude['%_COMPLETUDE'].apply(lambda x: f'{x:.1f}%'),
        textposition='outside',
        hovertemplate='<b>%{x}</b><br>Completude: %{y:.1f}%<extra></extra>'
    ))
    
    # Linha de meta
    fig_completude.add_hline(y=80, line_dash="dash", line_color="#ffd93d", annotation_text="Meta 80%")
    
    fig_completude.update_layout(
        height=400,
        xaxis_title="",
        yaxis_title="Completude (%)",
        yaxis=dict(range=[0, 110]),
        xaxis_tickangle=-45,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        margin=dict(l=50, r=20, t=20, b=100)
    )
    
    st.plotly_chart(fig_completude, use_container_width=True)

with col_right2:
    st.markdown("### 📊 Distribuição de Status dos Lotes")

    # Dados do donut - agrupando REAN + REPREAN em uma única categoria "Rep/Rean"
    total_rean = df_filtered['REAN'].sum()
    total_reprean = df_filtered['REPREAN'].sum()
    total_rep_rean = total_rean + total_reprean  # Agrupar repetição/reanálise

    status_data = {
        'Status': ['Aprovados', 'Reprovados', 'Rep/Rean'],
        'Quantidade': [total_aprovados, total_reprovados, total_rep_rean],
        'Cor': ['#00d4aa', '#ff6b6b', '#ffd93d']
    }
    df_status = pd.DataFrame(status_data)
    df_status = df_status[df_status['Quantidade'] > 0]  # Remover zeros
    
    fig_donut = go.Figure(data=[go.Pie(
        labels=df_status['Status'],
        values=df_status['Quantidade'],
        hole=0.5,
        marker_colors=df_status['Cor'],
        textinfo='label+percent',
        textposition='outside',
        hovertemplate='<b>%{label}</b><br>Quantidade: %{value:,.0f}<br>Percentual: %{percent}<extra></extra>'
    )])
    
    fig_donut.update_layout(
        height=400,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white'),
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=-0.2,
            xanchor="center",
            x=0.5
        ),
        margin=dict(l=20, r=20, t=20, b=60),
        annotations=[dict(
            text=f'<b>{total_lotes:,.0f}</b><br>Lotes',
            x=0.5, y=0.5,
            font_size=16,
            showarrow=False,
            font_color='white'
        )]
    )
    
    st.plotly_chart(fig_donut, use_container_width=True)

# ============================================================================
# TABELA COMPLETA DE TODOS OS ESTUDOS
# ============================================================================
st.markdown("---")
st.markdown("### 📋 Visão Completa de Todos os Estudos")
st.markdown("*Completude do pipeline analítico e status de aprovação*")

# Preparar tabela completa
df_tabela = df_filtered[df_filtered['TOTAL_LOTES'] > 0][[
    'EMPRESA_NOME', 'ESTUDO', 'ARQUIVO', 'TOTAL_LOTES', 
    'APROVADOS', '%_APROVADOS', 'REPROVADOS', '%_REPROVADOS',
    '%_ALIQUOTAGEM', '%_DOPAGEM', '%_EXTRACAO', '%_INJECAO', '%_BRUTOS', '%_COMPLETUDE'
]].copy()

df_tabela.columns = [
    'Empresa', 'Estudo', 'Arquivo', 'Lotes', 
    'Aprov.', '% Aprov.', 'Reprov.', '% Reprov.',
    '% Aliq.', '% Dop.', '% Extr.', '% Inj.', '% Brutos', '% Completude'
]

# Ordenar por completude
df_tabela = df_tabela.sort_values('% Completude', ascending=False)

st.dataframe(
    df_tabela,
    use_container_width=True,
    hide_index=True,
    height=400,
    column_config={
        "% Aprov.": st.column_config.ProgressColumn(
            "% Aprov.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Reprov.": st.column_config.ProgressColumn(
            "% Reprov.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Aliq.": st.column_config.ProgressColumn(
            "% Aliq.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Dop.": st.column_config.ProgressColumn(
            "% Dop.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Extr.": st.column_config.ProgressColumn(
            "% Extr.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Inj.": st.column_config.ProgressColumn(
            "% Inj.",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Brutos": st.column_config.ProgressColumn(
            "% Brutos",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
        "% Completude": st.column_config.ProgressColumn(
            "% Completude",
            format="%.1f%%",
            min_value=0,
            max_value=100,
        ),
    }
)

# ============================================================================
# TABELA DE ESTUDOS CRÍTICOS
# ============================================================================
st.markdown("---")
st.markdown("### ⚠️ Estudos que Requerem Atenção")
st.markdown("*Estudos com taxa de aprovação < 70% ou completude < 80%*")

df_criticos = df_filtered[
    ((df_filtered['%_APROVADOS'] < 70) | (df_filtered['%_COMPLETUDE'] < 80)) &
    (df_filtered['TOTAL_LOTES'] > 0)
][['EMPRESA_NOME', 'ESTUDO', 'ARQUIVO', 'TOTAL_LOTES', 'APROVADOS', '%_APROVADOS', 'REPROVADOS', '%_REPROVADOS', '%_COMPLETUDE']].copy()

df_criticos.columns = ['Empresa', 'Estudo', 'Arquivo', 'Total Lotes', 'Aprovados', '% Aprov.', 'Reprovados', '% Reprov.', '% Completude']

if len(df_criticos) > 0:
    # Ordenar por aprovação (menor primeiro)
    df_criticos = df_criticos.sort_values('% Aprov.', ascending=True).head(20)
    
    st.dataframe(
        df_criticos,
        use_container_width=True,
        hide_index=True,
        column_config={
            "% Aprov.": st.column_config.ProgressColumn(
                "% Aprov.",
                format="%.1f%%",
                min_value=0,
                max_value=100,
            ),
            "% Completude": st.column_config.ProgressColumn(
                "% Completude",
                format="%.1f%%",
                min_value=0,
                max_value=100,
            ),
        }
    )
else:
    st.success("✅ Nenhum estudo crítico encontrado com os filtros atuais!")

# ============================================================================
# EXPORTAÇÃO EXCEL ESTILIZADA
# ============================================================================
st.markdown("---")
st.markdown("### 📥 Exportar Dados")
st.markdown("*Exporte os dados filtrados em Excel com formatação profissional (2 abas)*")

def style_worksheet(ws, df_data, title, subtitle):
    """Aplica estilos a uma worksheet"""
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Cores para valores
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Título do relatório
    num_cols = len(df_data.columns)
    last_col = get_column_letter(max(num_cols, 1))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16, color="1e3a5f")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Subtítulo com data
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = subtitle
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Linha em branco
    ws.append([])
    
    # Adicionar dados
    for r_idx, row in enumerate(dataframe_to_rows(df_data, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Cabeçalho
            if r_idx == 4:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                # Colorir células de porcentagem baseado no valor
                col_name = df_data.columns[c_idx - 1] if c_idx <= len(df_data.columns) else ""
                if '%' in str(col_name) and isinstance(value, (int, float)):
                    if value >= 90:
                        cell.fill = verde
                    elif value >= 70:
                        cell.fill = amarelo
                    else:
                        cell.fill = vermelho
    
    # Auto-ajustar largura das colunas
    for col_idx, column in enumerate(df_data.columns, start=1):
        max_length = len(str(column)) + 2
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)
                except:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 25)
    
    # Colunas específicas mais largas
    if 'A' in ws.column_dimensions:
        ws.column_dimensions['A'].width = 18
    if 'B' in ws.column_dimensions:
        ws.column_dimensions['B'].width = 15
    if 'C' in ws.column_dimensions:
        ws.column_dimensions['C'].width = 35
    
    # Congelar painéis (cabeçalho sempre visível)
    ws.freeze_panes = 'A5'


def create_multi_sheet_excel(df_completo, df_criticos):
    """Cria um arquivo Excel com múltiplas abas estilizadas"""
    wb = Workbook()
    
    # Aba 1: Visão Completa
    ws1 = wb.active
    ws1.title = "Visão Completa"
    style_worksheet(
        ws1, 
        df_completo,
        "📋 Visão Completa de Todos os Estudos",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | {len(df_completo)} registros | Pipeline analítico e status de aprovação"
    )
    
    # Aba 2: Estudos Críticos
    ws2 = wb.create_sheet("Estudos Críticos")
    
    # Aplicar cor de aba (laranja para críticos)
    ws2.sheet_properties.tabColor = "FF6B6B"
    
    if len(df_criticos) > 0:
        style_worksheet(
            ws2, 
            df_criticos,
            "⚠️ Estudos que Requerem Atenção",
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | {len(df_criticos)} estudos críticos | Aprovação < 70% ou Completude < 80%"
        )
    else:
        ws2['A1'] = "✅ Nenhum estudo crítico encontrado com os filtros atuais!"
        ws2['A1'].font = Font(bold=True, size=14, color="00AA00")
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# Preparar dados para exportação - ABA 1: Visão Completa
df_export_completo = df_filtered[df_filtered['TOTAL_LOTES'] > 0][[
    'EMPRESA_NOME', 'ESTUDO', 'ARQUIVO', 'TOTAL_LOTES', 
    'APROVADOS', '%_APROVADOS', 'REPROVADOS', '%_REPROVADOS',
    '%_ALIQUOTAGEM', '%_DOPAGEM', '%_EXTRACAO', '%_INJECAO', '%_BRUTOS', '%_COMPLETUDE'
]].copy()

df_export_completo.columns = [
    'Empresa', 'Estudo', 'Arquivo', 'Total Lotes', 
    'Aprovados', '% Aprovados', 'Reprovados', '% Reprovados',
    '% Aliquotagem', '% Dopagem', '% Extração', '% Injeção', '% Brutos', '% Completude'
]

# Preparar dados para exportação - ABA 2: Estudos Críticos
df_export_criticos = df_filtered[
    ((df_filtered['%_APROVADOS'] < 70) | (df_filtered['%_COMPLETUDE'] < 80)) &
    (df_filtered['TOTAL_LOTES'] > 0)
][[
    'EMPRESA_NOME', 'ESTUDO', 'ARQUIVO', 'TOTAL_LOTES', 
    'APROVADOS', '%_APROVADOS', 'REPROVADOS', '%_REPROVADOS', '%_COMPLETUDE'
]].copy()

df_export_criticos.columns = [
    'Empresa', 'Estudo', 'Arquivo', 'Total Lotes', 
    'Aprovados', '% Aprovados', 'Reprovados', '% Reprovados', '% Completude'
]

# Ordenar por aprovação (menor primeiro)
df_export_criticos = df_export_criticos.sort_values('% Aprovados', ascending=True)

col_exp1, col_exp2, col_exp3 = st.columns([1, 1, 1])

with col_exp1:
    if OPENPYXL_AVAILABLE:
        excel_buffer = create_multi_sheet_excel(df_export_completo, df_export_criticos)
        st.download_button(
            label="📥 Baixar Excel (2 Abas)",
            data=excel_buffer,
            file_name=f"synvia_eln_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Excel com 2 abas: Visão Completa + Estudos Críticos"
        )
    else:
        st.warning("⚠️ Instale openpyxl para exportar Excel: pip install openpyxl")

with col_exp2:
    # CSV simples também
    csv_data = df_export_completo.to_csv(index=False, sep=';', encoding='utf-8-sig')
    st.download_button(
        label="📄 Baixar CSV",
        data=csv_data,
        file_name=f"synvia_eln_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
        help="Baixe os dados em formato CSV simples"
    )

with col_exp3:
    st.info(f"📊 **{len(df_export_completo)}** registros | ⚠️ **{len(df_export_criticos)}** críticos")

# ============================================================================
# FOOTER
# ============================================================================
st.markdown("---")
col_footer1, col_footer2, col_footer3, col_footer4 = st.columns(4)

with col_footer1:
    st.markdown(f"📊 **Registros filtrados:** {len(df_filtered)} de {len(df)}")

with col_footer2:
    source_icon = "☁️ SharePoint" if data_source == "sharepoint" else "📁 Local"
    st.markdown(f"📦 **Fonte:** {source_icon}")

with col_footer3:
    st.markdown(f"🔄 **Cache:** 5 minutos")

with col_footer4:
    st.markdown(f"⏰ **Acessado:** {datetime.now().strftime('%d/%m/%Y %H:%M')}")
